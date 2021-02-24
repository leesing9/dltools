from matplotlib import pyplot
import numpy
import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import os
import pandas
import seaborn
from pathlib import Path

header_worker = '작업자'
header_annotation = '미작업'
header_validation = '검수대기'
header_modification = '수정대기'
header_complete = '완료'
header_class = '라벨'

def adjustColumnWidth(workSheet, minimumRow, minimumColumn, maximumColumn):

    columnWidths = []
    for index, column in enumerate(workSheet.iter_cols(min_col = minimumColumn, max_col = maximumColumn, min_row = minimumRow)):
        for cell in column:
            value = cell.value
            if value is not None:
                if isinstance(value, str) is False:
                    value = str(value)
                try:
                    columnWidths[index] = max(columnWidths[index], len(value))
                except IndexError:
                    columnWidths.append(len(value))

def makePlot(dataFrame, figureSize, fontSize, saveFileName, x, y, hue = None):

    pyplot.rc('font', family = 'Malgun Gothic') # Font error

    pyplot.subplots(figsize = figureSize)
    plot = seaborn.barplot(x = x, y = y, hue = hue, data = dataFrame)
    plot.set(xlabel = None, ylabel = None)
    pyplot.xticks(fontsize = fontSize)

    showValuesOnBars(plot, 'h', 0.3)

    pyplot.savefig(saveFileName)

    return

def makeReport(dataFrame1, dataFrame2, saveExcelName, outdir):

    outdir = Path(outdir)
    modifiedDataFrame1 = pandas.melt(dataFrame1, id_vars=[header_worker], var_name='작업상태', value_name='개수')
    makePlot(dataFrame = modifiedDataFrame1, figureSize = (5, max([int(modifiedDataFrame1.shape[0] * 0.2),2])), fontSize = 10, saveFileName = str(outdir/'graph_dataFrame1.png'), x = '개수', y = header_worker, hue = '작업상태')

    modifiedDataFrame2 = pandas.melt(dataFrame2, id_vars = ['label'], var_name = '작업상태', value_name = '개수')
    makePlot(dataFrame = modifiedDataFrame2, figureSize = (5, max([int(modifiedDataFrame1.shape[0] * 0.2),2])), fontSize = 10, saveFileName = str(outdir/'graph_dataFrame2.png'), x = '개수', y = 'label', hue='작업상태')

    workBook = openpyxl.Workbook()

    workSheet = workBook.active
    workSheet.title = 'Report'

    for row in dataframe_to_rows(dataFrame1, index = False, header = True):

        workSheet.append(row)

    workSheet.append(['']) # Add empty row

    for row in dataframe_to_rows(dataFrame2, index = False, header = True):

        workSheet.append(row)

    rowsNumber1, columnsNumber1 = dataFrame1.shape
    dataFrameLastColumn1 = chr(ord('@') + columnsNumber1)

    adjustColumnWidth(workSheet, 1, 1, workSheet.max_column)
    setBorder(workSheet, f'A1:{dataFrameLastColumn1}{rowsNumber1 + 1}')

    rowsNumber2, columnsNumber2 = dataFrame2.shape
    dataFrameLastColumn2 = chr(ord('@') + columnsNumber2)

    # adjustColumnWidth(workSheet, 1, 1, workSheet.max_column)
    setBorder(workSheet, f'A{rowsNumber1 + 3}:{dataFrameLastColumn2}{rowsNumber1 + rowsNumber2 + 3}')
    print(f'A{rowsNumber1 + 3}:{dataFrameLastColumn2}{rowsNumber1 + rowsNumber2 + 3}')

    workSheet['G1'] = 'Amount of work per worker'
    workSheet.merge_cells('G1:M1')
    setBorder(workSheet, 'G1:M1')
    graph_sumPerTypeAndWorker = openpyxl.drawing.image.Image(str(outdir/'graph_dataFrame1.png'))
    workSheet.add_image(graph_sumPerTypeAndWorker, 'G2')

    workSheet['O1'] = 'Amount of work per class'
    workSheet.merge_cells('O1:U1')
    setBorder(workSheet, 'O1:U1')
    graph_sumPerType = openpyxl.drawing.image.Image(str(outdir/'graph_dataFrame2.png'))
    workSheet.add_image(graph_sumPerType, 'O2')

    workBook.save(str(outdir/f'{saveExcelName}.xlsx'))

    removeTemporaryFile(outdir)

    return

def removeTemporaryFile(outdir):

    os.remove(str(outdir/'graph_dataFrame1.png'))
    os.remove(str(outdir/'graph_dataFrame2.png'))

    return

def setBorder(workSheet, cellRange):

    thin = Side(border_style = 'thin', color = '000000')

    for row in workSheet[cellRange]:

        for cell in row:

            cell.border = Border(top = thin, left = thin, right = thin, bottom = thin)

def showValuesOnBars(graph, horizontalOrVertical = 'v', space = 0.4):

    def _showOnSinglePlot(plot):

        if horizontalOrVertical == 'v':

            for p in ax.patches:

                _x = p.get_x() + p.get_width() / 2
                _y = p.get_y() + p.get_height()
                value = int(p.get_height())
                ax.text(_x, _y, value, ha = 'center')

        elif horizontalOrVertical == 'h':

            for p in graph.patches:

                _x = p.get_x() + p.get_width() + float(space)
                _y = p.get_y() + p.get_height()
                value = int(p.get_width())
                plot.text(_x, _y, value, ha = 'left')

    if isinstance(graph, numpy.ndarray):

        for idx, ax in numpy.ndenumerate(graph):

            _showOnSinglePlot(graph)

    else:

        _showOnSinglePlot(graph)

    return

if __name__ == '__main__':

    # Set header (민규씨가 줄 DataFrame header)

    # DataFrame samples (카카오톡 내용으로 임의의 DataFrame 생성)
    dataFrame1 = pandas.DataFrame(data = {header_worker : ['미할당', '김씨', '박씨', '홍씨'], header_annotation : [0, 100, 200, 300], header_validation : [0, 100, 100, 200], header_modification : [0, 0, 100, 0], header_complete : [0, 100, 200, 100]})
    dataFrame2 = pandas.DataFrame(data = {header_class : ['결함', '박리', '박락'], header_annotation : [100, 200, 300], header_validation : [100, 100, 200], header_modification : [0, 100, 0], header_complete : [100, 200, 100]})

    makeReport(dataFrame1 = dataFrame1, dataFrame2 = dataFrame2, saveExcelName = 'Report')